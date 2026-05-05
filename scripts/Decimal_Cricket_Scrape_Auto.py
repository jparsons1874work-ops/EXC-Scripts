#!/usr/bin/env python3
"""
Decimal Cricket fixtures – scrape, compare strictly vs YESTERDAY, highlight, filter to today+, Slack-ready.
"""

import os, sys, io, re, time, argparse, datetime, glob, unicodedata
from typing import List, Optional
import pandas as pd

LOGIN   = "https://www.decimalcricket.net/login"
INFOAPP = "https://www.decimalcricket.net/info"

KEY_COLS = ["Start", "Name", "Format", "Competition", "Category"]

# ---------------- debug dump helpers ----------------
def save_state(driver, out_dir: str, tag: str):
    os.makedirs(out_dir, exist_ok=True)
    h = os.path.join(out_dir, f"{tag}.html")
    p = os.path.join(out_dir, f"{tag}.png")
    try:
        with open(h, "w", encoding="utf-8") as f:
            f.write(driver.page_source)
    except Exception:
        pass
    try:
        driver.save_screenshot(p)
    except Exception:
        pass
    print(f"[{tag}] url={getattr(driver, 'current_url', 'n/a')}\n       saved {h} & {p}")

# ---------------- parsing helpers ----------------
def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()

def pick_target_table(tables: List[pd.DataFrame]) -> Optional[pd.DataFrame]:
    want = [c.lower() for c in KEY_COLS]
    for df in tables:
        cols_norm = [_norm(c) for c in df.columns]
        if all(h in cols_norm for h in want):
            original = {_norm(c): c for c in df.columns}
            ordered  = [original[h] for h in want]
            extras   = [c for c in df.columns if _norm(c) not in want]
            return df[ordered + extras]
    return None

def read_tables_from_html_string(html: str) -> List[pd.DataFrame]:
    flavors = []
    try:
        import lxml  # noqa
        flavors.append("lxml")
    except Exception:
        pass
    try:
        import bs4, html5lib  # noqa
        flavors.append("bs4")
    except Exception:
        pass
    if not flavors:
        raise RuntimeError("Install parsers: pip install lxml beautifulsoup4 html5lib")
    last = None
    for flv in flavors:
        try:
            return pd.read_html(io.StringIO(html), flavor=flv)
        except Exception as e:
            last = e
    raise RuntimeError(f"pandas.read_html failed via {flavors}: {last}")

# -------- canonical key helpers --------
def to_dt_uk(s):
    dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
    try:
        return dt.dt.floor('min')
    except Exception:
        return pd.to_datetime(dt).floor('min') if pd.notna(dt) else dt

def canon_str(x: str) -> str:
    if x is None:
        return ""
    s = str(x)
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def build_key_frame(df: pd.DataFrame, key_cols: list) -> pd.Series:
    tmp = df.copy()
    if "Start" in tmp.columns:
        tmp["Start"] = to_dt_uk(tmp["Start"])
    for c in key_cols:
        if c != "Start" and c in tmp.columns:
            tmp[c] = tmp[c].map(canon_str)

    parts: List[pd.Series] = []
    for c in key_cols:
        if c == "Start":
            parts.append(tmp[c].astype("datetime64[ns]").dt.strftime("%Y-%m-%d %H:%M"))
        else:
            parts.append(tmp[c].astype(str))

    key = parts[0]
    for part in parts[1:]:
        key = key.str.cat(part, sep=" | ")
    return key

# ---------------- Excel formatting ----------------
def save_pretty_excel(df: pd.DataFrame, out_path: str, sheet_name="Fixtures"):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name=sheet_name)
        ws = xw.sheets[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="4F81BD")
        hdr_align = Alignment(vertical="center")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in ws[1]:
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border

        try:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, 1).number_format = "dd/mm/yyyy hh:mm"
        except Exception:
            pass

        for c_idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, ws.max_row), 1):
            mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(c_idx)].width = min(mx + 2, 60)

def save_combined_with_highlight(df_combined: pd.DataFrame, out_path: str, sheet_name="Combined"):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df_combined.to_excel(xw, index=False, sheet_name=sheet_name)
        ws = xw.sheets[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="4F81BD")
        hdr_align = Alignment(vertical="center")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in ws[1]:
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border

        if "_is_new" in df_combined.columns:
            new_fill = PatternFill("solid", fgColor="FFF2CC")
            headers = [c.value for c in ws[1]]
            new_idx = headers.index("_is_new") + 1 if "_is_new" in headers else None
            if new_idx:
                for r in range(2, ws.max_row + 1):
                    if ws.cell(r, new_idx).value is True:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(r, c).fill = new_fill

        try:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, 1).number_format = "dd/mm/yyyy hh:mm"
        except Exception:
            pass

        for c_idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, ws.max_row), 1):
            mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(c_idx)].width = min(mx + 2, 60)

# ---------------- Selenium (scraping) ----------------
def build_driver(headless: bool, profile_dir: str):
    """
    Build a Chrome driver WITHOUT webdriver_manager downloads at runtime.

    Priority:
      1) If CHROMEDRIVER_PATH is set and points to a file, use it.
      2) Otherwise, rely on Selenium Manager (Selenium 4.6+) / system PATH.
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service

    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1600,2400")

    if profile_dir:
        os.makedirs(profile_dir, exist_ok=True)
        opts.add_argument(f"--user-data-dir={os.path.abspath(profile_dir)}")

    chromedriver_path = os.getenv("CHROMEDRIVER_PATH", "").strip()
    if chromedriver_path:
        chromedriver_path = os.path.abspath(chromedriver_path)
        if not os.path.isfile(chromedriver_path):
            raise RuntimeError(f"CHROMEDRIVER_PATH is set but not a file: {chromedriver_path}")
        service = Service(executable_path=chromedriver_path)
        return webdriver.Chrome(service=service, options=opts)

    return webdriver.Chrome(options=opts)


def ensure_login(driver, username: str, password: str, out_dir: str):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    driver.get(LOGIN); time.sleep(0.3)
    save_state(driver, out_dir, "step_01_at_login_or_redirect")
    if "/login" not in driver.current_url.lower():
        save_state(driver, out_dir, "step_02_after_submit"); return
    email = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "input[type='email'],input[name='email'],input[name='username']")))
    pwd   = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "input[type='password'],input[name='password']")))
    email.clear(); email.send_keys(username)
    pwd.clear();   pwd.send_keys(password + Keys.ENTER)
    for _ in range(80):
        time.sleep(0.2)
        if "/login" not in driver.current_url.lower():
            break
    save_state(driver, out_dir, "step_02_after_submit")


def go_info_and_open_fixtures(driver, out_dir: str):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    driver.get(INFOAPP); time.sleep(0.5)
    save_state(driver, out_dir, "step_03_on_info_app")
    fixtures_pill = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, "//li[contains(@class,'nav-link') and (normalize-space()='FIXTURES' or @id='v-pills-d-tab')]"))
    )
    try:
        fixtures_pill.click()
    except Exception:
        driver.execute_script("arguments[0].click();", fixtures_pill)
    WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#v-pills-d")))

    def tab_has_table(drv):
        try:
            html = (drv.find_element(By.CSS_SELECTOR, "#v-pills-d").get_attribute("innerHTML") or "").lower()
            return "<table" in html or "tablesorter" in html or 'role="table"' in html
        except Exception:
            return False

    WebDriverWait(driver, 60).until(tab_has_table)
    save_state(driver, out_dir, "step_04_after_click_fixtures")


def get_fixtures_inner_html(driver) -> str:
    from selenium.webdriver.common.by import By
    return driver.find_element(By.CSS_SELECTOR, "#v-pills-d").get_attribute("innerHTML") or ""


def scrape_today_df(username: str, password: str, out_dir: str, profile_dir: str, headless: bool) -> pd.DataFrame:
    from selenium.common.exceptions import WebDriverException
    driver = None
    try:
        driver = build_driver(headless=headless, profile_dir=profile_dir)
        ensure_login(driver, username, password, out_dir)
        go_info_and_open_fixtures(driver, out_dir)
        html = get_fixtures_inner_html(driver)
        try:
            from io import StringIO
            tables = pd.read_html(StringIO(html), attrs={"id": "fixtures"})
            if not tables:
                tables = read_tables_from_html_string(html)
        except Exception:
            tables = read_tables_from_html_string(html)
        if not tables:
            raise RuntimeError("No tables parsed from Fixtures tab.")
        target = pick_target_table(tables)
        if target is None:
            target = tables[0]
        return target
    finally:
        try:
            if driver:
                driver.quit()
        except WebDriverException:
            pass

# --------- YESTERDAY file selection & read ---------
def _parse_date_from_name(path: str) -> Optional[datetime.date]:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
    if not m:
        return None
    try:
        return datetime.datetime.strptime(m.group(1), "%Y-%m-%d").date()
    except Exception:
        return None

def yesterday_file_path(out_dir: str, basename: str, today: datetime.date) -> Optional[str]:
    """Prefer *yesterday's* daily file; fallback to *yesterday's* combined; else latest daily before today."""
    y = today - datetime.timedelta(days=1)
    daily_y = os.path.join(out_dir, f"{basename}_{y:%Y-%m-%d}.xlsx")
    if os.path.exists(daily_y):
        return daily_y

    combined_y = os.path.join(out_dir, f"{basename}_combined_{y:%Y-%m-%d}.xlsx")
    if os.path.exists(combined_y):
        return combined_y

    # Fallback: most recent daily strictly before today
    cands = glob.glob(os.path.join(out_dir, f"{basename}_*.xlsx"))
    cands = [p for p in cands if "_combined_" not in os.path.basename(p)]
    cands = [(p, _parse_date_from_name(p)) for p in cands]
    cands = [p for p,d in cands if d is not None and d < today]
    cands.sort(key=_parse_date_from_name)
    return cands[-1] if cands else None

def read_prev_sheet_smart(path: str) -> pd.DataFrame:
    """Read Fixtures for daily; Combined for combined; else first sheet."""
    try:
        xl = pd.ExcelFile(path)
        name = os.path.basename(path)
        if "_combined_" in name and "Combined" in xl.sheet_names:
            return xl.parse("Combined")
        if "_combined_" not in name and "Fixtures" in xl.sheet_names:
            return xl.parse("Fixtures")
        return xl.parse(sheet_name=xl.sheet_names[0])
    except Exception:
        return pd.read_excel(path, sheet_name=0)

# -------- Merge / Highlight strictly vs yesterday --------
def combine_vs_yesterday(current_df: pd.DataFrame, prev_path: Optional[str]) -> pd.DataFrame:
    cur = current_df.copy()
    for c in KEY_COLS:
        if c not in cur.columns:
            raise RuntimeError(f"Expected column '{c}' not found in current data.")

    cur["Start"] = to_dt_uk(cur["Start"])
    cur["_key"]  = build_key_frame(cur, KEY_COLS)

    prev_keys = set()
    prev_full = None
    if prev_path and os.path.exists(prev_path):
        prev_full = read_prev_sheet_smart(prev_path)
        if "Start" in prev_full.columns:
            prev_full["Start"] = to_dt_uk(prev_full["Start"])
        prev_full["_key"] = build_key_frame(prev_full, KEY_COLS)
        prev_keys = set(prev_full["_key"].dropna().unique())

    cur["_is_new"] = ~cur["_key"].isin(prev_keys)

    # Combined = (yesterday for context, optional) ∪ current
    combo = pd.concat([prev_full, cur], ignore_index=True, sort=False) if prev_full is not None else cur
    combo = combo.drop_duplicates(subset=["_key"], keep="first").reset_index(drop=True)

    cols = KEY_COLS + [c for c in combo.columns if c not in KEY_COLS + ["_key", "_is_new"]]
    if "_is_new" in combo.columns:
        cols.append("_is_new")
    combo = combo[cols]
    return combo

# ---------------- Slack (optional) ----------------
def slack_upload(token: str, channel: str, paths: list, message: str = ""):
    try:
        from slack_sdk import WebClient
        from slack_sdk.errors import SlackApiError
    except Exception as e:
        print(f"Slack SDK not installed, skipping Slack upload: {e}", file=sys.stderr)
        return False
    client = WebClient(token=token)
    ok = True
    try:
        if message:
            client.chat_postMessage(channel=channel, text=message)
    except Exception as e:
        print(f"Slack message failed: {e}", file=sys.stderr)
        ok = False
    for p in paths:
        if not p or not os.path.exists(p):
            continue
        try:
            client.files_upload_v2(channel=channel, file=p, filename=os.path.basename(p), title=os.path.basename(p))
            print(f"Uploaded to Slack: {p}")
        except SlackApiError as e:
            print(f"Slack upload failed for {p}: {e.response.get('error')}", file=sys.stderr)
            ok = False
        except Exception as e:
            print(f"Slack upload failed for {p}: {e}", file=sys.stderr)
            ok = False
    return ok

# ---------------- main ----------------
def main():
    ap = argparse.ArgumentParser(description="Decimal Cricket fixtures – scrape, compare vs yesterday, highlight, filter to today+")
    ap.add_argument("--username", default=os.getenv("DECIMAL_USERNAME", ""))
    ap.add_argument("--password", default=os.getenv("DECIMAL_PASSWORD", ""))
    ap.add_argument("--out-dir", default=os.getenv("SCRIPT_OUTPUT_DIR", "runtime/output"))
    ap.add_argument("--basename", default="decimal_fixtures")
    ap.add_argument("--profile-dir", default=os.getenv("CHROME_PROFILE_DIR", "runtime/output/chrome_profile"))
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--slack-token", default=os.getenv("SLACK_BOT_TOKEN", ""))
    ap.add_argument("--slack-channel", default=os.getenv("SLACK_CHANNEL", ""))
    ap.add_argument("--slack-upload-log", default=os.getenv("SLACK_UPLOAD_LOG", "0"))
    args = ap.parse_args()

    if not args.username or not args.password:
        print("Set DECIMAL_USERNAME/DECIMAL_PASSWORD or pass --username/--password", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.out_dir, exist_ok=True)
    today = datetime.date.today()
    today_xlsx    = os.path.join(args.out_dir, f"{args.basename}_{today:%Y-%m-%d}.xlsx")
    combined_xlsx = os.path.join(args.out_dir, f"{args.basename}_combined_{today:%Y-%m-%d}.xlsx")
    rolling_xlsx  = os.path.join(args.out_dir, f"{args.basename}_combined_latest.xlsx")

    status_ok = True
    status_msg = ""
    try:
        # 1) Scrape today
        df_today = scrape_today_df(args.username, args.password, args.out_dir, args.profile_dir, args.headless)

        # Sort/normalize today's fixtures
        if "Start" in df_today.columns:
            df_today["Start"] = to_dt_uk(df_today["Start"])
            df_today = df_today.sort_values("Start", ascending=True).reset_index(drop=True)

        save_pretty_excel(df_today, today_xlsx, sheet_name="Fixtures")
        print(f"Saved today file: {today_xlsx}")

        # 2) Compare strictly vs YESTERDAY & highlight
        prev_path = yesterday_file_path(args.out_dir, args.basename, today)
        if prev_path:
            print(f"Yesterday file for comparison: {prev_path}")
        else:
            print("No yesterday/earlier daily file found; all rows considered new.")

        df_combined = combine_vs_yesterday(df_today, prev_path)

        # Keep only fixtures from TODAY forward
        if "Start" in df_combined.columns:
            df_combined["Start"] = pd.to_datetime(df_combined["Start"], errors="coerce")
            today_midnight = pd.Timestamp.today().normalize()
            df_combined = df_combined[df_combined["Start"] >= today_midnight].reset_index(drop=True)

        # Sort final
        if "Start" in df_combined.columns:
            df_combined = df_combined.sort_values("Start", ascending=True).reset_index(drop=True)

        save_combined_with_highlight(df_combined, combined_xlsx, sheet_name="Combined")
        print(f"Saved combined file: {combined_xlsx}")

        # Also refresh rolling (for convenience; not used for comparison anymore)
        try:
            save_combined_with_highlight(df_combined, rolling_xlsx, sheet_name="Combined")
            print(f"Updated rolling combined: {rolling_xlsx}")
        except Exception as e:
            print(f"Warning: failed to update rolling combined: {e}", file=sys.stderr)

        status_msg = f"✅ Decimal Cricket fixtures for {today:%Y-%m-%d}: new vs yesterday highlighted."
    except Exception as e:
        status_ok = False
        status_msg = f"❌ Decimal Cricket fixtures run failed for {today:%Y-%m-%d}. Error: {e}"
        print(status_msg, file=sys.stderr)

    # 3) Slack (optional)
    token = args.slack_token.strip()
    channel = args.slack_channel.strip()
    if token and channel:
        files_to_send = [combined_xlsx]
        if args.slack_upload_log.lower() in ("1", "true", "yes"):
            log_guess = os.path.join(args.out_dir, f"run_{today:%Y-%m-%d}.log")
            if os.path.exists(log_guess):
                files_to_send.append(log_guess)
        slack_upload(token, channel, files_to_send, status_msg)
    else:
        print("Slack not configured; skipping upload.")

    sys.exit(0 if status_ok else 2)

if __name__ == "__main__":
    main()
